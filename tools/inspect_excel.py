import pandas as pd
import json

file_path = '「2026」声度体验学员及报课名单.xlsx'

try:
    # 获取所有工作表名称
    # 使用 openpyxl 直接加载，不使用 pandas
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True)
    print(f"Sheets found: {wb.sheetnames}")
    
    for name in wb.sheetnames:
        print(f"\n--- Analyzing Sheet: {name} ---")
        ws = wb[name]
        # 获取第一行（表头）
        headers = [cell.value for cell in ws[1]]
        print(f"Headers: {headers}")
        
        # 获取第二行（数据示例）
        if ws.max_row > 1:
            first_row = [cell.value for cell in ws[2]]
            print(f"Example data: {first_row}")

except Exception as e:
    print(f"Error: {e}")
